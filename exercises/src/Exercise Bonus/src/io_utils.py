from __future__ import annotations

import io
import math
import zipfile
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

from config import CONFIG


TEMPLATE_MAP = {
    "uk_media": "BBC_news.xlsx",
    "uk_parliament": "UK_parliament.xlsx",
    "us_media": "NBC_news.xlsx",
    "us_congress": "US_Congress.xlsx",
}


def ensure_result_dirs() -> None:
    CONFIG.paths.results_excel.mkdir(parents=True, exist_ok=True)
    CONFIG.paths.results_graphs.mkdir(parents=True, exist_ok=True)
    CONFIG.paths.results_tables.mkdir(parents=True, exist_ok=True)
    CONFIG.paths.cache.mkdir(parents=True, exist_ok=True)


def _extract_template_bytes(template_name: str) -> bytes:
    with zipfile.ZipFile(CONFIG.paths.files_zip) as outer:
        for member in outer.namelist():
            if member.endswith(template_name) and "__MACOSX" not in member:
                return outer.read(member)
    raise FileNotFoundError(f"Template not found in files.zip: {template_name}")


def normalize_distribution(values: List[float], decimals: int | None = None) -> List[float]:
    decimals = decimals if decimals is not None else CONFIG.stage.decimal_places
    if not values:
        return values

    clean = [max(0.0, float(v)) for v in values]
    total = sum(clean)
    if total <= 0:
        equal = 1.0 / len(clean)
        rounded = [round(equal, decimals) for _ in clean]
    else:
        rounded = [round(v / total, decimals) for v in clean]

    drift = round(1.0 - sum(rounded), decimals)
    if abs(drift) >= 10 ** (-decimals):
        max_idx = max(range(len(rounded)), key=lambda i: rounded[i])
        rounded[max_idx] = round(rounded[max_idx] + drift, decimals)

    # handle any edge floating noise after correction
    final_total = round(sum(rounded), decimals)
    if not math.isclose(final_total, 1.0, rel_tol=0, abs_tol=10 ** (-decimals)):
        max_idx = max(range(len(rounded)), key=lambda i: rounded[i])
        rounded[max_idx] = round(rounded[max_idx] + (1.0 - final_total), decimals)
    return rounded


def export_channel_scores_to_template(
    channel: str,
    scores_df: pd.DataFrame,
    method: str,
) -> Path:
    template_bytes = _extract_template_bytes(TEMPLATE_MAP[channel])
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]

    topic_cols = [f"topic_{i}" for i in range(1, CONFIG.stage.topic_count + 1)]
    by_label: Dict[str, List[float]] = {}
    for _, row in scores_df.iterrows():
        by_label[row["time_label"]] = [float(row[c]) for c in topic_cols]

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if not isinstance(label, str) or not label.lower().startswith("time point"):
            continue
        values = by_label.get(label)
        if not values:
            values = [0.0] * CONFIG.stage.topic_count
        values = normalize_distribution(values)
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    file_name = f"{channel}_{method}.xlsx"
    out_path = CONFIG.paths.results_excel / file_name
    wb.save(out_path)
    return out_path
